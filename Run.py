import os
import shutil
import pandas as pd
from datetime import timedelta

# ===== 一開始先刪除 01~08 資料夾 =====
folders_all = ['01','02','03','04','05','06','07','08']
for f in folders_all:
    if os.path.exists(f):
        shutil.rmtree(f)
        print(f"Deleted folder: {f}")

# 設置檔案名稱
# inf1: 打卡紀錄
# inf2: 請假紀錄
inf1 = '1.xlsx'
inf2 = '2.xlsx'

# folder_mode 控制資料夾建立
# 1: 產生全部資料夾 (01~08)
# 2: 只保留最後 07、08，其餘執行完後刪除
folder_mode = 1
folders_all = ['01','02','03','04','05','06','07','08']
folders_final = ['07','08']

def create_folders(mode):
    if mode == 1:
        for f in folders_all:
            os.makedirs(f, exist_ok=True)
    elif mode == 2:
        for f in folders_final:
            os.makedirs(f, exist_ok=True)

create_folders(folder_mode)

# ===== Step 1: 分工號存檔 & 未滿8小時存檔 =====
df1 = pd.read_excel(inf1)
outdir1 = '01'
outdir2 = '02'

for name, group in df1.groupby('工號'):
    if folder_mode != 2:
        outpath1 = os.path.join(outdir1, f'{name}_all.xlsx')
        group.to_excel(outpath1, index=False)
        
        under8 = group[group['時數'] < 8]
        if not under8.empty:
            outpath2 = os.path.join(outdir2, f'{name}_under8.xlsx')
            under8.to_excel(outpath2, index=False)

# ===== Step 2: 分工號存請假資料 =====
df2 = pd.read_excel(inf2)
outdir3 = '03'

if folder_mode != 2:
    for name, group in df2.groupby('工號'):
        outpath1 = os.path.join(outdir3, f'{name}_leave.xlsx')
        group.to_excel(outpath1, index=False)

print("Step 1-2 Done!")

# ===== Step 3: 拆分期間欄位 =====
inf3 = "03"
outdir4 = "04"

if folder_mode != 2:
    for file in os.listdir(inf3):
        if not file.endswith(".xlsx"):
            continue
        file_path = os.path.join(inf3, file)
        df3 = pd.read_excel(file_path)
        
        if "期間" not in df3.columns:
            print(f"Warning: {file} not exist, Pass.")
            continue

        split_cols = df3["期間"].str.split("~", n=1, expand=True)
        df3["期間開始"] = pd.to_datetime(split_cols[0].str.strip(), errors="coerce")
        df3["期間結束"] = pd.to_datetime(split_cols[1].str.strip(), errors="coerce")
        df3 = df3.drop(columns=["期間"])

        cols = list(df3.columns)
        if "請假時數(hr)" in cols:
            idx = cols.index("請假時數(hr)")
            cols.insert(idx, cols.pop(cols.index("期間結束")))
            cols.insert(idx, cols.pop(cols.index("期間開始")))
            df3 = df3[cols]

        output_path = os.path.join(outdir4, f'{os.path.splitext(file)[0]}_split.xlsx')
        df3.to_excel(output_path, index=False)
        #print(f"Finish splitting：{file}")

print("Step 3 Done.")

# ===== Step 4: 時數大於8拆分每日8小時 =====
input_dir = "04"
output_dir = "05"
if folder_mode != 2:
    os.makedirs(output_dir, exist_ok=True)

for file in os.listdir(input_dir):
    if not file.endswith(".xlsx"):
        continue

    file_path = os.path.join(input_dir, file)
    df = pd.read_excel(file_path)
    new_rows = []

    hours_col = "請假時數(hr)"

    for idx, row in df.iterrows():
        total_hours = row.get(hours_col, 0)
        start_time = row.get("期間開始")

        if pd.isna(total_hours) or pd.isna(start_time):
            new_rows.append(row)
            continue

        start_time = pd.to_datetime(start_time)

        if total_hours <= 8:
            new_rows.append(row)
        else:
            full_days = int(total_hours // 8)
            remainder = total_hours % 8
            for i in range(full_days):
                new_row = row.copy()
                new_row["期間開始"] = start_time + timedelta(days=i)
                new_row["期間結束"] = row.get("類別", "X")
                new_row[hours_col] = 8
                new_rows.append(new_row)
            if remainder > 0:
                new_row = row.copy()
                new_row["期間開始"] = start_time + timedelta(days=full_days)
                new_row["期間結束"] = row.get("類別", "X")
                new_row[hours_col] = remainder
                new_rows.append(new_row)

    df_new = pd.DataFrame(new_rows)
    output_path = os.path.join(output_dir, f'{os.path.splitext(file)[0]}8.xlsx')
    df_new.to_excel(output_path, index=False)
    #print(f"Finish splitting hours：{file}")

print("Step 4 Done.")

# ===== Step 5: 合併未滿8小時 & 請假資料 =====
dir02 = "02"
dir05 = "05"
outdir06 = "06"
if folder_mode != 2:
    os.makedirs(outdir06, exist_ok=True)

for file02 in os.listdir(dir02):
    if not file02.endswith(".xlsx"):
        continue
    
    file_path02 = os.path.join(dir02, file02)
    df02 = pd.read_excel(file_path02)
    
    df02["歸屬日期"] = pd.to_datetime(df02["歸屬日期"]).dt.date
    hours_col02 = "時數"

    base_name = file02.split("_under8")[0]
    file05_name = f"{base_name}_leave_split8.xlsx"
    file05_path = os.path.join(dir05, file05_name)
    
    if os.path.exists(file05_path):
        df05 = pd.read_excel(file05_path)
        df05["日期"] = pd.to_datetime(df05["期間開始"]).dt.date
        df05 = df05[["日期", "請假時數(hr)", "類別"]].copy()
    else:
        df05 = pd.DataFrame(columns=["日期", "請假時數(hr)", "類別"])

    df02_idx = df02.set_index("歸屬日期")
    for idx, row in df05.iterrows():
        date = row["日期"]
        leave_hours = row["請假時數(hr)"]
        leave_type = row["類別"]
        if date in df02_idx.index:
            df02_idx.loc[date, hours_col02] += leave_hours
        else:
            new_row = {}
            for col in df02.columns:
                if col in ["工號", "姓名", "部門", "部門名稱"]:
                    new_row[col] = df02.iloc[0][col]
                elif col == hours_col02:
                    new_row[col] = leave_hours
                else:
                    new_row[col] = leave_type
            new_row["歸屬日期"] = date
            df02_idx = pd.concat([df02_idx, pd.DataFrame([new_row]).set_index("歸屬日期")])

    df_out = df02_idx.reset_index(drop=False)
    output_file = f"{os.path.splitext(file02)[0]}_add.xlsx"
    output_path = os.path.join(outdir06, output_file)
    df_out.to_excel(output_path, index=False)
    #print(f"Finish merging：{file02}")

print("Step 5 Done.")

# ===== Step 6~8: 過濾<8小時 & 統計總數量 =====
import openpyxl
from openpyxl.styles import Alignment

dir06 = "06"
outdir07 = "07"
outdir08 = "08"
os.makedirs(outdir07, exist_ok=True)
os.makedirs(outdir08, exist_ok=True)

summary_list = []

for file06 in os.listdir(dir06):
    if not file06.endswith("_add.xlsx"):
        continue

    file_path06 = os.path.join(dir06, file06)
    df = pd.read_excel(file_path06)
    df["歸屬日期"] = pd.to_datetime(df["歸屬日期"]).dt.date
    hours_col = "時數"

    df_det = df[df[hours_col] < 8].copy()
    keep_cols = ["工號", "姓名", "部門", "部門名稱", "歸屬日期", hours_col]
    other_cols = [c for c in df_det.columns if c not in keep_cols]
    df_det = df_det[keep_cols + other_cols]
    df_det = df_det.sort_values(by=["工號", "歸屬日期"])

    output_file_det = f"{os.path.splitext(file06)[0].replace('_add','')}_det.xlsx"
    output_path_det = os.path.join(outdir07, output_file_det)
    df_det.to_excel(output_path_det, index=False)
    #print(f"Finish filtering：{file06}")

    # ---- 對指定欄位靠右對齊 ----
    wb = openpyxl.load_workbook(output_path_det)
    ws = wb.active
    right_align_cols = ["最早刷卡日期","最早刷卡時間","最晚刷卡日期","最晚刷卡時間"]
    col_map = {cell.value: cell.column for cell in ws[1]}

    for col_name in right_align_cols:
        if col_name in col_map:
            col_letter = openpyxl.utils.get_column_letter(col_map[col_name])
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal='right')

    wb.save(output_path_det)

    if not df_det.empty:
        grouped = df_det.groupby(["工號", "姓名", "部門", "部門名稱"]).size().reset_index(name="總數量")
        summary_list.append(grouped)

if summary_list:
    df_summary = pd.concat(summary_list, ignore_index=True)
    df_summary = df_summary.groupby(["工號", "姓名", "部門", "部門名稱"], as_index=False)["總數量"].sum()
else:
    df_summary = pd.DataFrame(columns=["工號", "姓名", "部門", "部門名稱", "總數量"])

output_path_summary = os.path.join(outdir08, "finish.xlsx")
df_summary.to_excel(output_path_summary, index=False)

print("Step 6~8 Done.")
print("Finish merging all det files into finish.xlsx!")

# ---- 執行完畢後，如果 folder_mode = 2，刪除中間資料夾 ----
if folder_mode == 2:
    for f in set(folders_all) - set(folders_final):
        if os.path.exists(f):
            shutil.rmtree(f)
